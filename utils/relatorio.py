import json
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def carregar_json(caminho="database/producao.json"):
    if not os.path.exists(caminho):
        # Cria o arquivo vazio com lista
        with open(caminho, "w", encoding="utf-8") as f:
            json.dump([], f, ensure_ascii=False, indent=4)
        return []
    else:
        try:
            with open(caminho, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError:
            # Se o arquivo existir mas estiver vazio
            return []

class RelatorioExcel:
    def __init__(self, arquivo="relatorios/base_producao.xlsx"):
        self.arquivo = arquivo
        self.dados = carregar_json()  # Carrega todos os dados
        self.wb = None
        self.ws = None

    def criar_base(self):
        # Sempre cria um novo arquivo, apagando o anterior se existir
        self.wb = Workbook()
        self.ws = self.wb.active
        headers = [
            "Data", "Hora Início", "Hora Término", "Modelo", "OF", "Cliente",
            "Qtd Aprovada", "Motivo Falhas", "Qtd Retrabalhada",
            "Responsável", "Setor", "Turno"
        ]
        self.ws.append(headers)
        print("Nova base criada.")

    def preencher_planilha(self):
        if not self.ws:
            print("Planilha não carregada.")
            return

        for item in self.dados:
            self.ws.append([
                item.get("DATA"),
                item.get("HORA_INICIO"),
                item.get("HORA_TÉRMINO"),
                item.get("Modelo"),
                item.get("OF"),
                item.get("CLIENTE"),
                item.get("QTD_PEÇAS_APROVADAS"),
                item.get("MOTIVO_FALHAS"),
                item.get("QTD_RETRABALHADA"),
                item.get("RESPONSÁVEL"),
                item.get("setor"),
                item.get("TURNO")
            ])

    def formatar_planilha(self):
        # Define os estilos de alinhamento e fonte
        alinhamento_central = Alignment(horizontal='center', vertical='center')
        
        # Formata o cabeçalho (Linha 1)
        header_font = Font(bold=True, size=18)
        for cell in self.ws[1]:
            cell.font = header_font
            cell.alignment = alinhamento_central

        # Define a largura exata de cada coluna do cabeçalho
        self.ws.column_dimensions['A'].width = 12
        self.ws.column_dimensions['B'].width = 18
        self.ws.column_dimensions['C'].width = 22
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 11
        self.ws.column_dimensions['F'].width = 18
        self.ws.column_dimensions['G'].width = 24
        self.ws.column_dimensions['H'].width = 21
        self.ws.column_dimensions['I'].width = 26
        self.ws.column_dimensions['J'].width = 21
        self.ws.column_dimensions['K'].width = 15
        self.ws.column_dimensions['L'].width = 14

        # Formata o restante das células (a partir da linha 2)
        corpo_font = Font(size=11)
        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = corpo_font
                cell.alignment = alinhamento_central

    def salvar_relatorio(self):
        if not os.path.exists("relatorios"):
            os.makedirs("relatorios")
        self.wb.save(self.arquivo)
        print(f"Base salva como {self.arquivo}")